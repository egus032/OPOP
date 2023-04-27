import glob
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

import xlsxwriter

counter = 0

g_path = r'C:\Users\User\Desktop\Рабочая папка\от Геращенковой\2022\c 01.08.22\Нагрузка\Таблицы для рассылки\*'

folder_with_files = glob.glob(g_path + '\*.xlsx')

# border = Border(left=Side(border_style=None, color='FF000000'), right=Side(border_style=None, color='FF000000'), top=Side(border_style=None, color='FF000000'),
#                 bottom=Side(border_style=None, color='FF000000'), diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0, outline=Side(border_style=None,
#         color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))

# top = Side(border_style='dashed', color='FF0707')

list_with_values = []

for g_file in folder_with_files:
    workbook = openpyxl.load_workbook(g_file)
    ws = workbook['Таблица_для кадровой_справки']
    #   Установить цвет фона для столбцов
    ws.column_dimensions['M'].fill = PatternFill("solid", start_color="5cb800")
    ws.column_dimensions['N'].fill = PatternFill("solid", start_color="5cb800")
    #   Установка ширины для столбцов
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 34
    ws.column_dimensions['H'].width = 34
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 34
    ws.column_dimensions['L'].width = 34
    ws.column_dimensions['M'].width = 45
    ws.column_dimensions['N'].width = 45
    ws.column_dimensions['O'].width = 34
    ws.column_dimensions['P'].width = 24
    ws.column_dimensions['Q'].width = 24
    ws.column_dimensions['R'].width = 24
    
    counter += 1
    print(counter, g_file)
    workbook.save(g_file)
    
    
        

    
    