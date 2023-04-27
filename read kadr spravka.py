import openpyxl
import pandas as pd
import glob

g_folder = glob.glob(r'C:\Users\User\Desktop\Рабочая папка\от Геращенковой\2022\c 01.08.22\Таблица для формирования по кадрам\*.xlsx')

# for g_file in g_folder:
#     g_sheet = openpyxl.load_workbook(g_file)
#     print(f'{g_sheet.sheetnames, g_file}')
#     # print(g_sheet.sheetnames, g_file)